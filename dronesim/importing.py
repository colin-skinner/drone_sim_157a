from openpyxl import load_workbook
import pandas as pd
import os

class ThrustData:

    def __init__(self, relative_path: str, *, drop_duplicates = True):

        base_path = os.getcwd()

        workbook = load_workbook(f"{base_path}/{relative_path}")
        sheet = workbook["Sheet1"]

        self.applied_voltage = 14.8

        # Loading "Applied Voltage" and "Throttle Range"
        for column_num in range(1,sheet.max_column):
            cell = sheet.cell(1, column_num).value

            if "Applied Voltage" in str(cell):
                voltage = cell.split(" ")[-2]
                self.applied_voltage = float(voltage)

            if "Throttle Range" in str(cell):
                throttle_range_str = cell.split(":")[1]
                throttle_range_str = throttle_range_str.split(" ")[:-1]
                
                values: list[int] = []

                for value in throttle_range_str:
                    try:
                        value = int(value)
                        values.append(value)
                    except ValueError:
                        continue
                
                self.min_throttle = min(values)
                self.max_throttle = max(values)
                self.applied_voltage = float(voltage)

        # Finding table
        header_row = None
        for row_num in range(1,sheet.max_row):
            cell = sheet.cell(row_num, 1).value
            if "Time" in str(cell):
                header_row = row_num


        if header_row is None:
            raise ImportError("No time in .xlsx")
        
        # Making table
        data = list(sheet.values)[header_row-1:]
        self.lookup_table = pd.DataFrame(data[1:], columns=data[0])

        if drop_duplicates:
            self.lookup_table = self.lookup_table.drop_duplicates(subset=["ESC signal (µs)"])


class TrajectoryData: 

    def __init__(self, relative_path: str, sheet_name: str, extra_cols: list[str] | None = None):
        base_path = os.getcwd()

        workbook = load_workbook(f"{base_path}/{relative_path}")
        sheet = workbook[sheet_name].values

        columns = next(sheet)
        sheet_data = pd.DataFrame(sheet, columns=columns)

        if extra_cols is None:
            extra_cols = []
        
        self.state_df = sheet_data[["t",
                                    "r_x", "r_y", "r_z",
                                    "v_x", "v_y", "v_z",
                                    "a_x", "a_y", "a_z",
                                    "omega_x", "omega_y", "omega_z",
                                    "n_x", "n_y", "n_z", "theta"] + extra_cols]

        self.data: dict[list, tuple[list, list]] = {}
        for i, row in self.state_df.iterrows():
            row = [float(i) for i in row]
            row[3] = row[3]
            self.data |= {row[0]: (row[1:4], row[4:7])}

        
        
    

        





