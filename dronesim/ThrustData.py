from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os, sys

class ThrustData:

    def __init__(self, relative_path: str, *, drop_duplicates = False):

        base_path = os.getcwd()

        workbook = load_workbook(f"{base_path}/{relative_path}")
        sheet = workbook["Sheet1"]

        self.applied_voltage = 14.8

        # Loading "Applied Voltage" and "Throttle Range"
        for column_num in range(1,sheet.max_column):
            cell = sheet.cell(1, column_num).value

            if "Applied Voltage" in str(cell):
                voltage = cell.split(" ")[-2]
                self.applied_voltage = float(voltage)

            if "Throttle Range" in str(cell):
                throttle_range_str = cell.split(":")[1]
                throttle_range_str = throttle_range_str.split(" ")[:-1]
                
                values: list[int] = []

                for value in throttle_range_str:
                    try:
                        value = int(value)
                        values.append(value)
                    except ValueError:
                        continue
                
                self.min_throttle = min(values)
                self.max_throttle = max(values)
                self.applied_voltage = float(voltage)

        # Finding table
        header_row = None
        for row_num in range(1,sheet.max_row):
            cell = sheet.cell(row_num, 1).value
            if "Time" in str(cell):
                header_row = row_num


        if header_row is None:
            raise ImportError("No time in .xlsx")
        
        # Making table
        data = list(sheet.values)[header_row-1:]
        self.lookup_table = pd.DataFrame(data[1:], columns=data[0])

        if drop_duplicates:
            self.lookup_table = self.lookup_table.drop_duplicates(subset=['ESC signal (µs)'])

        


